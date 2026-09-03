#!/usr/bin/env python3
"""
QA hậu-kiểm cho file '<mã> - kstudy import - ....xlsx' (đã sinh, hoặc đã sửa tay
trước khi import vào /admin/curriculum). Build script đã QA course.json; script này
để kiểm lại CHÍNH file xlsx.

Kiểm: code 6 ký tự ALL CAPS; ai_context 150-250 từ; prerequisites không trống; curriculum_md
không trống; mỗi lesson content 300-600 từ; Radar đúng 6 trục; mọi skill_tag trong content thuộc
Radar tags; không có marker runtime cấm ([BÀI TẬP]/[ĐIỂM:]/[QUIZ_INLINE]/[QUIZ_RESULT]/[LEVEL_UP]/
[COMPLETE_LEVEL]/[CHOICES]) trong Info/Lessons/Resources — marker này chỉ hệ thống sinh
(xem COURSE_CONTENT_GUIDELINE.md §1), nhúng sẵn vào content DB là FAIL.

Dùng: python validate_course.py "<file import.xlsx>"   (exit 1 nếu có FAIL)
"""
import sys, re
import openpyxl

def wc(s): return len((s or "").split())
res = []
def log(s, m): res.append((s, m))

FORBIDDEN_MARKERS = ["[BÀI TẬP]", "[BAI TAP]", "[QUIZ_INLINE]", "[QUIZ_RESULT]", "[LEVEL_UP]", "[CHOICES]"]
FORBIDDEN_MARKER_PREFIXES = ["[ĐIỂM:", "[DIEM:", "[COMPLETE_LEVEL"]

def find_markers(text):
    t = (text or "").upper()
    return [mk for mk in FORBIDDEN_MARKERS if mk in t] + [mk for mk in FORBIDDEN_MARKER_PREFIXES if mk in t]

def main():
    if len(sys.argv) < 2:
        print('Dùng: python validate_course.py "<import.xlsx>"'); sys.exit(2)
    wb = openpyxl.load_workbook(sys.argv[1])
    need = {"Info", "Lessons", "Resources", "Sessions", "Assignments", "Radar"}
    miss = need - set(wb.sheetnames)
    if miss:
        log("FAIL", f"Thiếu sheet: {', '.join(sorted(miss))} (không khớp template)")
    # Info
    info = {}
    if "Info" in wb.sheetnames:
        ws = wb["Info"]
        for r in range(2, ws.max_row + 1):
            k = ws.cell(r, 1).value
            if k: info[str(k).strip()] = ws.cell(r, 2).value
    code = str(info.get("code", "") or "")
    log("PASS" if re.fullmatch(r'[A-Z0-9]{6}', code) else "FAIL",
        f"code '{code}'" + ("" if re.fullmatch(r'[A-Z0-9]{6}', code) else " (cần 6 ký tự A-Z/0-9 ALL CAPS)"))
    n = wc(info.get("ai_context", ""))
    log("PASS" if 150 <= n <= 250 else ("WARN" if n < 150 else "FAIL"), f"ai_context {n} từ (150-250)")
    prereq = str(info.get("prerequisites", "") or "").strip()
    log("PASS" if prereq else "WARN",
        "prerequisites: đã có điều kiện tiên quyết" if prereq else "prerequisites: trống — xác nhận lại với người tạo khóa (LUÔN phải có, kể cả khi là 'không điều kiện')")
    curr = str(info.get("curriculum_md", "") or "").strip()
    log("PASS" if curr else "FAIL",
        "curriculum_md: đã có tổng quan + lộ trình" if curr else "curriculum_md: trống — courses.curriculum_md sẽ rỗng (CẤM theo checklist publish)")
    # Radar
    radar_tags, axes = set(), 0
    if "Radar" in wb.sheetnames:
        ws = wb["Radar"]
        for r in range(2, ws.max_row + 1):
            if not ws.cell(r, 1).value: continue
            axes += 1
            for t in re.split(r'[,\s]+', str(ws.cell(r, 3).value or "")):
                t = t.strip()
                if t: radar_tags.add(t)
    log("PASS" if axes == 6 else "FAIL", f"Radar {axes} trục (cần 6)")
    # Lessons
    used = set()
    marker_hits = []
    for lbl, txt in ([("Info.curriculum_md", curr), ("Info.ai_context", info.get("ai_context", ""))]):
        hits = find_markers(txt)
        if hits: marker_hits.append((lbl, hits))
    if "Lessons" in wb.sheetnames:
        ws = wb["Lessons"]
        for r in range(2, ws.max_row + 1):
            so = ws.cell(r, 1).value
            content = ws.cell(r, 4).value or ""
            if so is None and not content: continue
            body = "\n".join(l for l in content.splitlines() if not l.lower().strip().startswith("skill_tags"))
            n = wc(body)
            log("PASS" if 300 <= n <= 600 else ("WARN" if n < 300 else "FAIL"), f"Bài {so} content {n} từ (300-600)")
            for line in re.findall(r'skill_tags?\s*:\s*(.+)', content, re.I):
                for t in re.split(r'[,\s]+', re.split(r'[.;]', line)[0]):
                    t = t.strip()
                    if t: used.add(t)
            hits = find_markers(content)
            if hits: marker_hits.append((f"Bài {so} content", hits))
    if "Resources" in wb.sheetnames:
        ws = wb["Resources"]
        for r in range(2, ws.max_row + 1):
            desc = ws.cell(r, 5).value or ""
            hits = find_markers(desc)
            if hits: marker_hits.append((f"Resource dòng {r}", hits))
    orphan = sorted(used - radar_tags)
    log("FAIL" if orphan else "PASS",
        f"skill_tag lệch radar: {', '.join(orphan)}" if orphan else f"{len(used)} skill_tag đều khớp radar")
    if marker_hits:
        for lbl, hits in marker_hits:
            log("FAIL", f"{lbl}: chứa marker runtime cấm {hits}")
    else:
        log("PASS", "Không có marker runtime cấm trong Info/Lessons/Resources")

    order = {"FAIL": 0, "WARN": 1, "PASS": 2}
    nf = sum(1 for s, _ in res if s == "FAIL")
    nw = sum(1 for s, _ in res if s == "WARN")
    print("=== QA import.xlsx ===")
    for s, m in sorted(res, key=lambda x: order[x[0]]):
        print(f"[{s}] {m}")
    print(f"\n{nf} FAIL · {nw} WARN · {len(res)-nf-nw} PASS")
    sys.exit(1 if nf else 0)

if __name__ == "__main__":
    main()
